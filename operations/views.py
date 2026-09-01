import csv
from io import BytesIO
from zipfile import BadZipFile
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Count, Max, Q, Sum
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.datetime import to_excel
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.exceptions import InvalidFileException
from .access import module_required
from .forms import (BulkProgramForm, CloseProductionForm, ProcessMovementForm,
                    ProgramOrderForm, StartProductionForm, SurplusMovementForm)
from .models import (Employee, Inventory, InventoryBucket, Machine, Movement, Process, ProductionClose,
                     ProductionOrder, WorkInProcess)
from .services import (close_production, create_program_order, move_process_material,
                       move_surplus, resolve_program_client, start_production)


@login_required
def dashboard(request):
    context = {
        "open_orders": ProductionOrder.objects.filter(status=ProductionOrder.Status.OPEN).count(),
        "active_work": WorkInProcess.objects.filter(status=WorkInProcess.Status.ACTIVE).count(),
        "closed_count": ProductionClose.objects.count(),
        "part_count": Inventory.objects.count(),
        "recent_closes": ProductionClose.objects.select_related("work_item__order__part", "work_item__machine")[:8],
    }
    return render(request, "operations/dashboard.html", context)


@login_required
@module_required("program_loading")
def order_list(request):
    orders = ProductionOrder.objects.select_related("part", "part__client").order_by("-created_at")[:500]
    return render(request, "operations/order_list.html", {"orders": orders})


@login_required
@module_required("program_loading")
def load_program(request):
    form = ProgramOrderForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        employee = Employee.objects.get(payroll_number=form.cleaned_data["payroll_number"])
        order = create_program_order(
            client_name=form.cleaned_data["client"], part_number=form.cleaned_data["part_number"],
            program=form.cleaned_data["program"], quantity=form.cleaned_data["quantity"],
            required_date=form.cleaned_data["required_date"], line=form.cleaned_data["line"],
            comment=form.cleaned_data["comment"], employee=employee, user=request.user)
        messages.success(request, f"Programa cargado correctamente con el folio {order.folio}.")
        return redirect("order-list")
    return render(request, "operations/program_load.html", {"form": form, "active_tab": "single"})


def _cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_quantity(value):
    # El archivo original tiene cantidades numéricas con formato de fecha.
    # openpyxl las convierte a datetime; el serial recupera el número capturado.
    if isinstance(value, (datetime, date, time)):
        return Decimal(str(to_excel(value)))
    return Decimal(str(value).replace(",", ""))


@login_required
@module_required("program_loading")
def download_program_template(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    headers = ["ID Cliente", "Orden de Produccion", "Linea Prod Clte",
               "Fecha de Entrega", "Num. Parte", "Cantidad", "Linea"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="0875BD")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:G1000"
    widths = {"A": 18, "B": 25, "C": 22, "D": 20, "E": 22, "F": 15, "G": 18}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in range(2, 1001):
        sheet.cell(row, 4).number_format = "dd/mm/yyyy"
        sheet.cell(row, 6).number_format = "0.###"

    positive_quantity = DataValidation(
        type="decimal", operator="greaterThan", formula1="0", allow_blank=True)
    positive_quantity.error = "La cantidad debe ser un número mayor que cero."
    positive_quantity.errorTitle = "Cantidad inválida"
    positive_quantity.prompt = "Captura una cantidad mayor que cero."
    positive_quantity.promptTitle = "Cantidad"
    positive_quantity.showErrorMessage = True
    positive_quantity.showInputMessage = True
    sheet.add_data_validation(positive_quantity)
    positive_quantity.add("F2:F1000")

    instructions = workbook.create_sheet("Instrucciones")
    instructions.column_dimensions["A"].width = 28
    instructions.column_dimensions["B"].width = 85
    instructions.append(["Campo", "Descripción"])
    descriptions = [
        ("ID Cliente", "Identificador o nombre del cliente. Obligatorio."),
        ("Orden de Produccion", "Número del programa u orden del cliente. Obligatorio."),
        ("Linea Prod Clte", "Línea de producción del cliente. Opcional."),
        ("Fecha de Entrega", "Fecha en formato dd/mm/aaaa. Opcional."),
        ("Num. Parte", "Número de parte. Obligatorio."),
        ("Cantidad", "Cantidad numérica mayor que cero. Obligatorio."),
        ("Linea", "Línea interna asignada al programa. Opcional; tiene prioridad sobre Linea Prod Clte."),
    ]
    for description in descriptions:
        instructions.append(description)
    for cell in instructions[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    instructions.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_carga_programas.xlsx"'
    return response


@login_required
@module_required("program_loading")
def download_completed_programs(request):
    orders = ProductionOrder.objects.filter(
        Q(status=ProductionOrder.Status.COMPLETE) | Q(remaining_quantity=0)
    ).select_related("part", "part__client").annotate(
        completed_quantity=Sum("work_items__closes__quantity"),
        last_close=Max("work_items__closes__closed_at"),
    ).order_by("program", "part__client__name", "part__number")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Programas completados"
    headers = [
        "Semana / Orden de Produccion", "Folio", "ID Cliente", "Cliente",
        "Num. Parte", "Cantidad Programada", "Cantidad Terminada", "Restante",
        "Fecha de Entrega", "Linea", "Estado", "Ultimo Cierre",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="0875BD")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for order in orders:
        client = order.part.client
        completed = order.completed_quantity
        if completed is None:
            completed = order.quantity - order.remaining_quantity
        last_close = order.last_close
        if last_close and timezone.is_aware(last_close):
            last_close = timezone.localtime(last_close).replace(tzinfo=None)
        sheet.append([
            order.program, order.folio, client.external_id if client else "",
            client.name if client else "", order.part.number, order.quantity,
            completed, order.remaining_quantity, order.required_date, order.line,
            order.get_status_display(), last_close,
        ])

    widths = [30, 22, 16, 26, 22, 21, 20, 14, 18, 18, 16, 22]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:L{max(sheet.max_row, 1)}"
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 9).number_format = "dd/mm/yyyy"
        sheet.cell(row, 12).number_format = "dd/mm/yyyy hh:mm"
        for column in (6, 7, 8):
            sheet.cell(row, column).number_format = "0.###"

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = timezone.localdate().strftime("programas_completados_%Y-%m-%d.xlsx")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@module_required("program_loading")
def bulk_load_program(request):
    form = BulkProgramForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        try:
            workbook = load_workbook(form.cleaned_data["file"], data_only=True, read_only=True)
            sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
            expected = ["ID Cliente", "Orden de Produccion", "Linea Prod Clte",
                        "Fecha de Entrega", "Num. Parte", "Cantidad"]
            actual = [_cell_text(sheet.cell(1, col).value) for col in range(1, 7)]
            line_header = _cell_text(sheet.cell(1, 7).value)
            if actual != expected or line_header not in ("", "Linea"):
                raise ValidationError(
                    "Los encabezados no coinciden con la plantilla. Se esperan: "
                    + ", ".join(expected + ["Linea"]))
            employee = Employee.objects.get(payroll_number=form.cleaned_data["payroll_number"])
            rows, errors = [], []
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                client, program, customer_line = map(_cell_text, row[:3])
                required, part_number, raw_quantity = row[3], _cell_text(row[4]), row[5]
                line = _cell_text(row[6] if len(row) > 6 else "") or customer_line
                if not any((client, program, part_number, raw_quantity)):
                    continue
                try:
                    quantity = _cell_quantity(raw_quantity)
                except (InvalidOperation, TypeError, ValueError):
                    errors.append(f"Fila {row_number}: cantidad inválida.")
                    continue
                if not client or not program or not part_number or quantity <= 0:
                    errors.append(f"Fila {row_number}: cliente, orden, parte y cantidad son obligatorios.")
                    continue
                if isinstance(required, datetime):
                    required = required.date()
                elif not isinstance(required, date):
                    required = None
                rows.append((client, program, line, required, part_number, quantity))
            if errors:
                raise ValidationError(errors[:12])
            if not rows:
                raise ValidationError("El archivo no contiene filas válidas para importar.")
            with transaction.atomic():
                for client, program, line, required, part_number, quantity in rows:
                    resolved_client = resolve_program_client(
                        client_reference=client, part_number=part_number)
                    create_program_order(client_name=client, part_number=part_number,
                                         program=program, quantity=quantity,
                                         required_date=required, line=line,
                                         employee=employee, user=request.user,
                                         comment=f"Carga masiva: {form.cleaned_data['file'].name}",
                                         client=resolved_client)
        except (ValidationError, KeyError, ValueError, BadZipFile, InvalidFileException) as exc:
            form.add_error("file", exc)
        else:
            messages.success(request, f"Carga masiva terminada: {len(rows)} programas agregados.")
            return redirect("order-list")
    return render(request, "operations/program_load.html", {"form": form, "active_tab": "bulk"})


@login_required
@module_required("surplus")
def surplus(request):
    form = SurplusMovementForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        employee = Employee.objects.get(payroll_number=form.cleaned_data["payroll_number"])
        try:
            move_surplus(part=form.cleaned_data["part"], action=form.cleaned_data["action"],
                         quantity=form.cleaned_data["quantity"], employee=employee,
                         program=form.cleaned_data["program"], comment=form.cleaned_data["comment"],
                         user=request.user)
        except ValidationError as exc:
            form.add_error(None, exc)
        else:
            messages.success(request, "Movimiento de material sobrante registrado.")
            return redirect("inventory-list")
    recent = Movement.objects.filter(movement_type=Movement.Type.SURPLUS).select_related("part", "employee").order_by("-occurred_at")[:12]
    return render(request, "operations/surplus.html", {"form": form, "recent": recent})


@login_required
@module_required("process_material")
def process_list(request):
    processes = Process.objects.filter(active=True).prefetch_related("machine_set")
    return render(request, "operations/process_list.html", {"processes": processes})


@login_required
@module_required("process_material")
def process_material(request):
    form = ProcessMovementForm(request.POST or None)
    selected_part = None
    if request.method == "POST" and request.POST.get("part"):
        selected_part = form.fields["part"].queryset.filter(pk=request.POST["part"]).first()
    if request.method == "POST" and form.is_valid():
        selected_part = form.cleaned_data["part"]
        employee = Employee.objects.get(payroll_number=form.cleaned_data["payroll_number"])
        try:
            move_process_material(
                part=selected_part, source_process=form.cleaned_data["source_process"],
                destination_process=form.cleaned_data["destination_process"],
                program=form.cleaned_data["program"], quantity=form.cleaned_data["quantity"],
                employee=employee, comment=form.cleaned_data["comment"], user=request.user)
        except ValidationError as exc:
            form.add_error(None, exc)
        else:
            messages.success(request, "El material fue transferido al siguiente proceso.")
            return redirect(f"{request.path}?part={selected_part.pk}")
    part_id = request.GET.get("part")
    if part_id and not selected_part:
        selected_part = form.fields["part"].queryset.filter(pk=part_id).first()
    inventory = None
    buckets = []
    programs = []
    process_total = 0
    if selected_part:
        inventory = Inventory.objects.filter(part=selected_part).first()
        if inventory:
            buckets = inventory.buckets.filter(kind="PROCESS").order_by("name")
            programs = inventory.buckets.filter(kind="PROGRAM", quantity__gt=0).order_by("name")
            process_total = buckets.aggregate(total=Sum("quantity"))["total"] or 0
    recent = Movement.objects.filter(movement_type=Movement.Type.INVENTORY).select_related(
        "part", "employee").order_by("-occurred_at")[:15]
    return render(request, "operations/process_material.html", {
        "form": form, "selected_part": selected_part, "inventory": inventory,
        "buckets": buckets, "programs": programs, "process_total": process_total,
        "recent": recent})


@login_required
@module_required("heliang")
def work_list(request):
    items = WorkInProcess.objects.select_related("order__part", "machine").order_by("-started_at")[:500]
    return render(request, "operations/work_list.html", {"items": items})


@login_required
@module_required("heliang")
def heliang(request):
    action = request.POST.get("action")
    selected_order = None
    if request.method == "GET" and request.GET.get("order"):
        selected_order = ProductionOrder.objects.filter(
            pk=request.GET["order"], status=ProductionOrder.Status.OPEN
        ).first()
    start_initial = ({"order": selected_order, "quantity": selected_order.remaining_quantity}
                     if selected_order else None)
    start_form = StartProductionForm(
        request.POST if action == "start" else None,
        prefix="start", initial=start_initial,
    )
    selected_work = None
    if request.method == "GET" and request.GET.get("work"):
        selected_work = WorkInProcess.objects.filter(
            pk=request.GET["work"], status=WorkInProcess.Status.ACTIVE
        ).select_related("order__part", "machine").first()
    close_initial = ({"work_item": selected_work, "quantity": selected_work.remaining_quantity}
                     if selected_work else None)
    close_form = CloseProductionForm(
        request.POST if action == "close" else None,
        prefix="close", initial=close_initial,
    )
    if request.method == "POST" and action == "start" and start_form.is_valid():
        employee = Employee.objects.get(payroll_number=start_form.cleaned_data["payroll_number"])
        try:
            work = start_production(
                order=start_form.cleaned_data["order"], machine=start_form.cleaned_data["machine"],
                quantity=start_form.cleaned_data["quantity"], employee=employee, user=request.user)
        except ValidationError as exc:
            start_form.add_error(None, exc)
        else:
            messages.success(request, f"Orden {work.order.folio} asignada a {work.machine.code} con folio {work.folio}.")
            return redirect("heliang")
    if request.method == "POST" and action == "close" and close_form.is_valid():
        employee = Employee.objects.get(payroll_number=close_form.cleaned_data["payroll_number"])
        try:
            close = close_production(
                work_item=close_form.cleaned_data["work_item"],
                quantity=close_form.cleaned_data["quantity"], employee=employee,
                user=request.user, comment=close_form.cleaned_data["comment"])
        except ValidationError as exc:
            close_form.add_error(None, exc)
        else:
            messages.success(request, f"Producción cerrada con el folio {close.folio}.")
            return redirect("heliang")
    open_orders = ProductionOrder.objects.filter(status=ProductionOrder.Status.OPEN).select_related(
        "part", "part__client").order_by("required_date", "created_at")[:100]
    active_items = WorkInProcess.objects.filter(status=WorkInProcess.Status.ACTIVE).select_related(
        "order__part", "machine", "started_by").order_by("started_at")
    recent_closes = ProductionClose.objects.select_related(
        "work_item__order__part", "work_item__machine", "closed_by").order_by("-closed_at")[:30]
    occupied_ids = set(active_items.values_list("machine_id", flat=True))
    machines = Machine.objects.filter(active=True).select_related("process").order_by("code")
    machine_rows = [{"machine": machine, "occupied": machine.pk in occupied_ids}
                    for machine in machines]
    return render(request, "operations/heliang.html", {
        "start_form": start_form, "close_form": close_form, "open_orders": open_orders,
        "active_items": active_items, "recent_closes": recent_closes,
        "machine_rows": machine_rows, "selected_order": selected_order,
        "selected_work": selected_work})


@login_required
@module_required("line_dashboard")
def line_dashboard(request):
    now = timezone.localtime()
    today = now.date()
    active_items = list(WorkInProcess.objects.filter(
        status=WorkInProcess.Status.ACTIVE).select_related(
        "order__part", "order__part__client", "machine", "started_by"))
    active_by_machine = {item.machine_id: item for item in active_items}
    machine_rows = []
    for machine in Machine.objects.filter(active=True).order_by("code"):
        work = active_by_machine.get(machine.pk)
        progress = 0
        if work and work.initial_quantity:
            progress = min(100, round(float(
                (work.initial_quantity - work.remaining_quantity) / work.initial_quantity * 100)))
        machine_rows.append({"machine": machine, "work": work, "progress": progress})
    today_closes = ProductionClose.objects.filter(closed_at__date=today)
    today_summary = today_closes.aggregate(quantity=Sum("quantity"), weight=Sum("weight_kg"))
    process_totals = InventoryBucket.objects.filter(kind="PROCESS", quantity__gt=0).values(
        "name").annotate(total=Sum("quantity")).order_by("-total")[:10]
    open_orders = ProductionOrder.objects.filter(status=ProductionOrder.Status.OPEN).select_related(
        "part", "part__client").order_by("required_date", "created_at")[:8]
    recent_closes = ProductionClose.objects.filter(closed_at__date=today).select_related(
        "work_item__order__part", "work_item__machine").order_by("-closed_at")[:8]
    total_machines = len(machine_rows)
    return render(request, "operations/line_dashboard.html", {
        "now": now, "machine_rows": machine_rows, "active_count": len(active_items),
        "available_count": total_machines - len(active_items), "total_machines": total_machines,
        "open_count": ProductionOrder.objects.filter(status=ProductionOrder.Status.OPEN).count(),
        "today_closes_count": today_closes.count(),
        "today_quantity": today_summary["quantity"] or 0,
        "today_weight": today_summary["weight"] or 0,
        "process_totals": process_totals, "open_orders": open_orders,
        "recent_closes": recent_closes,
    })


@login_required
@module_required("line_dashboard")
def progress_dashboard(request):
    orders = ProductionOrder.objects.exclude(
        status=ProductionOrder.Status.CANCELLED
    ).select_related("part", "part__client").annotate(
        completed_quantity=Sum("work_items__closes__quantity")
    ).order_by("program", "part__client__name", "part__number")

    progress_data = []
    for order in orders:
        completed = order.completed_quantity or Decimal("0")
        progress_data.append({
            "folio": order.folio,
            "week": order.program or "Sin semana",
            "client": order.part.client.name if order.part.client else "Sin cliente",
            "part": order.part.number,
            "programmed": float(order.quantity),
            "completed": float(completed),
        })
    return render(request, "operations/progress_dashboard.html", {
        "progress_data": progress_data,
    })


@login_required
@module_required("heliang")
def start_work(request):
    form = StartProductionForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        employee = Employee.objects.get(payroll_number=form.cleaned_data["payroll_number"])
        try:
            work = start_production(order=form.cleaned_data["order"], machine=form.cleaned_data["machine"],
                quantity=form.cleaned_data["quantity"], employee=employee, user=request.user)
        except ValidationError as exc:
            form.add_error(None, exc)
        else:
            messages.success(request, f"La orden pasó a producción con el folio {work.folio}.")
            return redirect("work-list")
    return render(request, "operations/form.html", {"form": form, "title": "Enviar a producción", "button": "Iniciar producción"})


@login_required
@module_required("heliang")
def close_work(request):
    form = CloseProductionForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        employee = Employee.objects.get(payroll_number=form.cleaned_data["payroll_number"])
        try:
            close = close_production(work_item=form.cleaned_data["work_item"], quantity=form.cleaned_data["quantity"],
                employee=employee, user=request.user, comment=form.cleaned_data["comment"])
        except ValidationError as exc:
            form.add_error(None, exc)
        else:
            messages.success(request, f"Producción registrada con el folio {close.folio}.")
            return redirect("report")
    return render(request, "operations/form.html", {"form": form, "title": "Cerrar producción", "button": "Registrar cierre"})


@login_required
@module_required("inventory")
def inventory_list(request):
    query = request.GET.get("q", "").strip()
    rows = Inventory.objects.select_related("part", "part__client").prefetch_related("buckets")
    if query: rows = rows.filter(part__number__icontains=query)
    return render(request, "operations/inventory_list.html", {"rows": rows[:300], "query": query})


@login_required
@module_required("reports")
def report(request):
    rows = ProductionClose.objects.select_related("work_item__order__part", "work_item__machine", "closed_by")
    start, end = request.GET.get("start"), request.GET.get("end")
    if start: rows = rows.filter(closed_at__date__gte=start)
    if end: rows = rows.filter(closed_at__date__lte=end)
    return render(request, "operations/report.html", {"rows": rows[:1000], "start": start, "end": end})


@login_required
@module_required("reports")
def report_csv(request):
    rows = ProductionClose.objects.select_related("work_item__order__part", "work_item__machine", "closed_by")
    start, end = request.GET.get("start"), request.GET.get("end")
    if start: rows = rows.filter(closed_at__date__gte=start)
    if end: rows = rows.filter(closed_at__date__lte=end)
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="reporte_produccion.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(["Folio", "Orden", "Programa", "Número de parte", "Cantidad", "Máquina", "Cierre", "Turno", "Kilogramos", "Operador", "Comentario"])
    for row in rows:
        writer.writerow([row.folio, row.work_item.order.folio, row.work_item.order.program, row.work_item.order.part.number,
            row.quantity, row.work_item.machine.code, row.closed_at, row.shift, row.weight_kg, row.closed_by or "", row.comment])
    return response
