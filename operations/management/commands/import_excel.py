from collections import Counter
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel
from operations.models import (Client, Employee, Inventory, InventoryBucket, Machine, Movement, Part, Process,
                               ProductionClose, ProductionOrder, WorkInProcess)


PROCESS_NAMES = ["Corte", "Doblez", "Expansion", "Reduccion", "Perforacion", "Perforacion Con Broca",
                 "Indentacion", "Beading", "Extrusion", "Dimple", "Cepillado", "Corte Con Broca",
                 "Rectificado", "Ovalamiento", "Soldadura", "Empaque"]

def text(value):
    if value is None: return ""
    if isinstance(value, float) and value.is_integer(): return str(int(value))
    return str(value).strip()

def number(value):
    if value in (None, ""): return Decimal("0")
    # Algunas plantillas tienen cantidades numéricas con estilo de fecha.
    # openpyxl las entrega como datetime; el serial conserva la cantidad original.
    if isinstance(value, (datetime, date, time)):
        try: return Decimal(str(to_excel(value)))
        except (TypeError, ValueError): return Decimal("0")
    try: return Decimal(str(value).replace(",", ""))
    except (InvalidOperation, ValueError): return Decimal("0")

def as_date(value):
    if isinstance(value, datetime): return value.date()
    if isinstance(value, date): return value
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d%m%Y"):
        try: return datetime.strptime(text(value), fmt).date()
        except ValueError: pass
    return None

def as_datetime(date_value, time_value=None):
    day = as_date(date_value) or timezone.localdate()
    clock = time(0, 0)
    if isinstance(time_value, datetime): clock = time_value.time()
    elif isinstance(time_value, time): clock = time_value
    elif isinstance(time_value, (float, int)):
        seconds = round(float(time_value) * 86400) % 86400
        clock = time(seconds // 3600, (seconds % 3600) // 60, seconds % 60)
    elif text(time_value):
        for fmt in ("%H:%M:%S", "%H:%M"):
            try: clock = datetime.strptime(text(time_value), fmt).time(); break
            except ValueError: pass
    return timezone.make_aware(datetime.combine(day, clock), timezone.get_current_timezone())


class Command(BaseCommand):
    help = "Importa y concilia los datos históricos del XLSM de MESA."
    def add_arguments(self, parser):
        parser.add_argument("workbook", type=Path)
        parser.add_argument("--orders", type=Path, help="Plantilla Sheet1.xlsx opcional (solo validación).")
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        source = options["workbook"]
        if not source.exists(): raise CommandError(f"No existe {source}")
        stats = Counter()
        with transaction.atomic():
            wb = load_workbook(source, data_only=True, read_only=True, keep_vba=True)
            required = {"RESUMEN", "USUARIOS", "PROGRAMAS_HELIANS", "PROCESANDO_HELIANS", "CERRADOS_HELIANS", "PESOS"}
            missing = required - set(wb.sheetnames)
            if missing: raise CommandError(f"Faltan hojas requeridas: {', '.join(sorted(missing))}")
            self.import_catalogs(wb, stats)
            self.import_inventory(wb, stats)
            self.import_orders(wb, stats)
            self.import_work(wb, stats)
            self.import_closes(wb, stats)
            self.import_movements(wb, stats)
            self.import_admin_accounts(wb, stats)
            if options.get("orders"): self.import_order_template(options["orders"], stats)
            if options["dry_run"]:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING("Modo simulación: no se guardaron cambios."))
        self.stdout.write(self.style.SUCCESS("Importación terminada"))
        for key, value in sorted(stats.items()): self.stdout.write(f"  {key}: {value}")

    def employee(self, payroll):
        payroll = text(payroll)
        return Employee.objects.filter(payroll_number=payroll).first() if payroll else None

    def part(self, part_number, client_name=""):
        part_number = text(part_number)
        if not part_number: return None
        client = None
        if text(client_name):
            code = text(client_name).upper()[:30]
            client, _ = Client.objects.get_or_create(code=code, defaults={"name": text(client_name)})
        obj, _ = Part.objects.get_or_create(number=part_number, defaults={"client": client})
        if client and not obj.client_id: obj.client = client; obj.save(update_fields=["client", "updated_at"])
        return obj

    def import_catalogs(self, wb, stats):
        for row in wb["USUARIOS"].iter_rows(min_row=2, values_only=True):
            payroll = text(row[0])
            if not payroll: continue
            Employee.objects.update_or_create(payroll_number=payroll, defaults={"name": text(row[1]) or payroll,
                "area": text(row[2]), "identifier": text(row[3]), "active": True})
            stats["empleados"] += 1
        for index, name in enumerate(PROCESS_NAMES, 1):
            Process.objects.update_or_create(code=name.lower().replace(" ", "-"), defaults={"name": name, "position": index})
        if "PROCESOS" in wb.sheetnames:
            for row in wb["PROCESOS"].iter_rows(min_row=2, values_only=True):
                code = text(row[5] if len(row) > 5 else "")
                if code: Machine.objects.get_or_create(code=code)
        for row in wb["PESOS"].iter_rows(min_row=2, values_only=True):
            part_number = text(row[0])
            if not part_number: continue
            obj = self.part(part_number, row[3] if len(row) > 3 else "")
            weight = number(row[2] if len(row) > 2 else 0)
            if weight: obj.unit_weight_kg = weight; obj.description = text(row[1]); obj.save()
            stats["pesos"] += 1

    def import_inventory(self, wb, stats):
        ws = wb["RESUMEN"]
        program_names = [text(ws.cell(3, col).value) or f"Programa {col-4}" for col in range(5, 15)]
        pending_buckets = {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            part_number = text(row[2] if len(row) > 2 else "")
            if not part_number: continue
            part = self.part(part_number, row[1] if len(row) > 1 else "")
            inv, _ = Inventory.objects.update_or_create(part=part, defaults={
                "surplus": number(row[31] if len(row) > 31 else 0), "real": number(row[32] if len(row) > 32 else 0)})
            for offset, name in enumerate(program_names, 4):
                bucket = InventoryBucket(inventory=inv, kind="PROGRAM", name=name,
                    quantity=number(row[offset] if len(row) > offset else 0))
                pending_buckets[(inv.pk, bucket.kind, bucket.name)] = bucket
            for offset, name in enumerate(PROCESS_NAMES, 14):
                bucket = InventoryBucket(inventory=inv, kind="PROCESS", name=name,
                    quantity=number(row[offset] if len(row) > offset else 0))
                pending_buckets[(inv.pk, bucket.kind, bucket.name)] = bucket
            stats["inventarios"] += 1
            if len(pending_buckets) >= 5000:
                InventoryBucket.objects.bulk_create(list(pending_buckets.values()), update_conflicts=True,
                    update_fields=["quantity", "updated_at"], unique_fields=["inventory", "kind", "name"])
                pending_buckets.clear()
        if pending_buckets:
            InventoryBucket.objects.bulk_create(list(pending_buckets.values()), update_conflicts=True,
                update_fields=["quantity", "updated_at"], unique_fields=["inventory", "kind", "name"])

    def import_orders(self, wb, stats):
        for row in wb["PROGRAMAS_HELIANS"].iter_rows(min_row=2, values_only=True):
            folio, program, part_number = text(row[0]), text(row[1]), text(row[2])
            if not folio or not part_number: continue
            qty = number(row[3]); status_text = text(row[5]).lower()
            ProductionOrder.objects.update_or_create(folio=folio, defaults={"program": program, "part": self.part(part_number),
                "quantity": qty, "remaining_quantity": qty, "required_date": as_date(row[4]), "line": text(row[7]),
                "status": ProductionOrder.Status.OPEN if status_text not in ("cerrado", "cancelado") else ProductionOrder.Status.COMPLETE,
                "loaded_by": self.employee(row[8]), "legacy_id": folio})
            stats["ordenes_abiertas"] += 1

    def ensure_order(self, folio, program, part_number, quantity=0):
        obj = ProductionOrder.objects.filter(folio=folio).first()
        if obj: return obj
        qty = number(quantity)
        return ProductionOrder.objects.create(folio=folio or f"LEGACY-O-{ProductionOrder.objects.count()+1}", program=text(program),
            part=self.part(part_number), quantity=qty, remaining_quantity=0, status=ProductionOrder.Status.COMPLETE, legacy_id=text(folio))

    def import_work(self, wb, stats):
        for row in wb["PROCESANDO_HELIANS"].iter_rows(min_row=2, values_only=True):
            folio = text(row[0]); part_number = text(row[3])
            if not folio or not part_number: continue
            qty = number(row[4]); machine, _ = Machine.objects.get_or_create(code=text(row[7]) or "SIN-MAQUINA")
            order = self.ensure_order(text(row[1]), row[2], part_number, qty)
            WorkInProcess.objects.update_or_create(folio=folio, defaults={"order": order, "machine": machine,
                "initial_quantity": qty, "remaining_quantity": qty, "status": WorkInProcess.Status.ACTIVE,
                "started_at": as_datetime(row[5], row[6]), "started_by": self.employee(row[9]), "legacy_id": folio})
            stats["ordenes_procesando"] += 1

    def import_closes(self, wb, stats):
        for row in wb["CERRADOS_HELIANS"].iter_rows(min_row=2, values_only=True):
            folio, work_folio, order_folio, part_number = text(row[0]), text(row[1]), text(row[2]), text(row[4])
            if not folio or not part_number: continue
            qty = number(row[5]); order = self.ensure_order(order_folio, row[3], part_number, qty)
            work = WorkInProcess.objects.filter(folio=work_folio).first()
            if not work:
                machine, _ = Machine.objects.get_or_create(code=text(row[10]) or "SIN-MAQUINA")
                work = WorkInProcess.objects.create(folio=work_folio or f"LEGACY-P-{WorkInProcess.objects.count()+1}",
                    order=order, machine=machine, initial_quantity=qty, remaining_quantity=0,
                    status=WorkInProcess.Status.CLOSED, started_at=as_datetime(row[6], row[7]), legacy_id=work_folio)
            ProductionClose.objects.update_or_create(folio=folio, defaults={"work_item": work, "quantity": qty,
                "weight_kg": number(row[17] if len(row) > 17 else 0), "closed_at": as_datetime(row[8], row[9]),
                "shift": text(row[12]), "closed_by": self.employee(row[13]), "comment": text(row[16]), "legacy_id": folio})
            stats["cierres"] += 1

    def import_movements(self, wb, stats):
        mappings = [("MOVIMIENTOS", Movement.Type.INVENTORY), ("MOVIMIENTO DE PROGRAMAS", Movement.Type.PROGRAM)]
        for sheet, movement_type in mappings:
            if sheet not in wb.sheetnames: continue
            for row in wb[sheet].iter_rows(min_row=2, values_only=True):
                if not text(row[0]): continue
                if movement_type == Movement.Type.INVENTORY:
                    Movement.objects.get_or_create(legacy_source=f"{sheet}:{text(row[0])}", defaults={"folio": text(row[0]),
                        "movement_type": movement_type, "part": self.part(row[2], row[1]), "source": text(row[3]),
                        "destination": text(row[4]), "quantity": number(row[5]), "employee": self.employee(row[6]),
                        "occurred_at": as_datetime(row[9], row[10]), "comment": text(row[11]), "program": text(row[12])})
                else:
                    Movement.objects.get_or_create(legacy_source=f"{sheet}:{text(row[0])}", defaults={"folio": text(row[0]),
                        "movement_type": movement_type, "part": self.part(row[2], row[1]), "destination": text(row[7]),
                        "quantity": number(row[3]), "employee": self.employee(row[4]), "occurred_at": as_datetime(row[8], row[9]),
                        "comment": text(row[10])})
                stats["movimientos"] += 1

    def import_admin_accounts(self, wb, stats):
        if "ADMINISTRADORES" not in wb.sheetnames: return
        User = get_user_model()
        for row in wb["ADMINISTRADORES"].iter_rows(min_row=2, values_only=True):
            username = text(row[0])
            if not username: continue
            user, created = User.objects.get_or_create(username=username, defaults={"is_staff": True})
            if created: user.set_unusable_password(); user.save(update_fields=["password"])
            stats["administradores_sin_contrasena"] += 1

    def import_order_template(self, source, stats):
        if not source.exists(): raise CommandError(f"No existe {source}")
        wb = load_workbook(source, data_only=True, read_only=True)
        if "Sheet1" not in wb.sheetnames: raise CommandError("La plantilla no contiene la hoja Sheet1")
        expected = ["ID Cliente", "Orden de Produccion", "Linea Prod Clte", "Fecha de Entrega", "Num. Parte", "Cantidad"]
        actual = [text(wb["Sheet1"].cell(1, col).value) for col in range(1, 7)]
        if actual != expected: raise CommandError(f"Encabezados inesperados: {actual}")
        client_names = {}
        if "Sheet2" in wb.sheetnames:
            client_names = {text(row[0]): text(row[1]) for row in wb["Sheet2"].iter_rows(min_row=2, values_only=True) if text(row[0])}
        date_key = timezone.localdate().strftime("%d%m%Y")
        sequence = ProductionOrder.objects.filter(folio__startswith=f"O{date_key}-").count()
        for row_number, row in enumerate(wb["Sheet1"].iter_rows(min_row=2, values_only=True), 2):
            client_code, program, required, part_number, qty = text(row[0]), text(row[1]), as_date(row[3]), text(row[4]), number(row[5])
            if not program or not part_number or qty <= 0: continue
            legacy_id = f"Sheet1.xlsx:{row_number}"
            existing = ProductionOrder.objects.filter(legacy_id=legacy_id).first()
            client_name = client_names.get(client_code, client_code)
            part = self.part(part_number, client_name)
            line = text(row[6] if len(row) > 6 else "") or text(row[2] if len(row) > 2 else "")
            if existing:
                existing.program, existing.part, existing.quantity = program, part, qty
                existing.required_date, existing.line = required, line
                existing.save()
            else:
                sequence += 1
                ProductionOrder.objects.create(folio=f"O{date_key}-{sequence}", program=program, part=part,
                    quantity=qty, remaining_quantity=qty, required_date=required, line=line,
                    status=ProductionOrder.Status.OPEN, legacy_id=legacy_id)
            stats["ordenes_desde_plantilla"] += 1
