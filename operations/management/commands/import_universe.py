import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from operations.models import Client, Part


def text(value):
    return "" if value is None else str(value).strip()


def client_code(name):
    return re.sub(r"[^A-Z0-9]+", "-", name.upper()).strip("-")[:30]


class Command(BaseCommand):
    help = "Importa la relación cliente/número de parte desde Universo Ramos."

    def add_arguments(self, parser):
        parser.add_argument("workbook", type=Path)

    @transaction.atomic
    def handle(self, *args, **options):
        source = options["workbook"]
        if not source.exists():
            raise CommandError(f"No existe {source}")

        workbook = load_workbook(source, data_only=True, read_only=True)
        sheet_name = "Sheet1 (2)" if "Sheet1 (2)" in workbook.sheetnames else "Sheet1"
        sheet = workbook[sheet_name]
        has_external_id = sheet_name == "Sheet1 (2)"
        updated_parts = 0
        missing_parts = 0
        clients_seen = set()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = text(row[1] if len(row) > 1 else "")
            external_id = text(row[2] if has_external_id and len(row) > 2 else "")
            part_index = 4 if has_external_id else 3
            part_number = text(row[part_index] if len(row) > part_index else "")
            if not name or name.upper() == "CLIENTE" or not part_number or part_number.upper() == "NÚMERO DE PARTE":
                continue

            if external_id.upper() in {"#N/A", "N/A", "NA"}:
                external_id = ""
            code = client_code(name)
            client, _ = Client.objects.get_or_create(code=code, defaults={"name": name})
            changed = []
            if client.name != name:
                client.name = name
                changed.append("name")
            if client.external_id != external_id:
                client.external_id = external_id
                changed.append("external_id")
            if changed:
                changed.append("updated_at")
                client.save(update_fields=changed)
            clients_seen.add(client.pk)

            part = Part.objects.filter(number=part_number).first()
            if part is None:
                missing_parts += 1
                continue
            if part.client_id != client.pk:
                part.client = client
                part.save(update_fields=["client", "updated_at"])
            updated_parts += 1

        self.stdout.write(self.style.SUCCESS("Relación de clientes importada"))
        self.stdout.write(f"  clientes: {len(clients_seen)}")
        self.stdout.write(f"  números de parte relacionados: {updated_parts}")
        self.stdout.write(f"  números de parte no encontrados: {missing_parts}")
