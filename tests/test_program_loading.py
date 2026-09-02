from decimal import Decimal
from io import BytesIO
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from operations.models import (Client, Inventory, InventoryBucket, Machine, Movement,
                               Part, Process, ProductionClose, ProductionOrder, WorkInProcess)
from operations.services import (create_program_order, move_process_material, move_surplus,
                                 resolve_program_client)


class ProgramLoadingTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user("tester", password="test")
        self.employee = self.user

    def test_individual_load_creates_order_and_traceability(self):
        order = create_program_order(client_name="Cliente Uno", part_number="P-100",
                                     program="OP-55", quantity="25", employee=self.employee,
                                     comment="Urgente", user=self.user)
        self.assertEqual(order.remaining_quantity, Decimal("25"))
        self.assertEqual(order.part.client.name, "Cliente Uno")
        self.assertTrue(Movement.objects.filter(folio=order.folio,
                                                movement_type=Movement.Type.PROGRAM).exists())

    def test_downloaded_bulk_template_has_required_structure(self):
        self.user.is_superuser = True
        self.user.save(update_fields=["is_superuser"])
        self.client.force_login(self.user)

        response = self.client.get(reverse("download-program-template"))

        self.assertEqual(response.status_code, 200)
        self.assertIn("plantilla_carga_programas.xlsx", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Sheet1", "Instrucciones"])
        headers = [workbook["Sheet1"].cell(1, column).value for column in range(1, 8)]
        self.assertEqual(headers, ["ID Cliente", "Orden de Produccion", "Linea Prod Clte",
                                   "Fecha de Entrega", "Num. Parte", "Cantidad", "Linea"])

    def test_completed_programs_can_be_downloaded_as_excel(self):
        self.user.is_superuser = True
        self.user.save(update_fields=["is_superuser"])
        self.client.force_login(self.user)
        customer = Client.objects.create(
            code="LENNOX-1", name="LENNOX 1", external_id="CD0036")
        part = Part.objects.create(number="P-CLOSED", client=customer)
        closed_order = ProductionOrder.objects.create(
            folio="O-CLOSED", program="S35", part=part,
            quantity=100, remaining_quantity=0,
            status=ProductionOrder.Status.COMPLETE, line="L1",
        )
        machine = Machine.objects.create(code="M-CLOSED")
        work = WorkInProcess.objects.create(
            folio="P-CLOSED", order=closed_order, machine=machine,
            initial_quantity=100, remaining_quantity=0,
            status=WorkInProcess.Status.CLOSED, started_at=timezone.now(),
        )
        ProductionClose.objects.create(
            folio="C-CLOSED", work_item=work, quantity=100,
            closed_at=timezone.now(),
        )
        ProductionOrder.objects.create(
            folio="O-OPEN", program="S36", part=part,
            quantity=50, remaining_quantity=50,
            status=ProductionOrder.Status.OPEN,
        )

        response = self.client.get(reverse("download-completed-programs"))

        self.assertEqual(response.status_code, 200)
        self.assertIn("programas_completados_", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook["Programas completados"]
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual([sheet.cell(2, column).value for column in range(1, 9)], [
            "S35", "O-CLOSED", "CD0036", "LENNOX 1", "P-CLOSED", 100, 100, 0,
        ])

    def test_bulk_load_uses_original_line_column(self):
        self.user.is_superuser = True
        self.user.save(update_fields=["is_superuser"])
        self.client.force_login(self.user)
        expected_client = Client.objects.create(code="CLIENTE-UNO", name="Cliente Uno", external_id="ID001")
        Part.objects.create(number="P-200", client=expected_client)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["ID Cliente", "Orden de Produccion", "Linea Prod Clte",
                      "Fecha de Entrega", "Num. Parte", "Cantidad", "Linea"])
        sheet.append(["ID001", "OP-200", "Cliente-L1", None, "P-200", 600, "LINEA-7"])
        sheet["F2"].number_format = "mm-dd-yy"
        output = BytesIO()
        workbook.save(output)
        upload = SimpleUploadedFile(
            "programas.xlsx", output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(reverse("bulk-load-program"), {
            "file": upload,
        })

        self.assertRedirects(response, reverse("order-list"))
        order = ProductionOrder.objects.get(program="OP-200")
        self.assertEqual(order.line, "LINEA-7")
        self.assertEqual(order.quantity, Decimal("600"))
        self.assertEqual(order.part.client, expected_client)

    def test_client_resolution_falls_back_to_part_when_id_is_missing(self):
        expected_client = Client.objects.create(code="RHEEM-NIPLES", name="RHEEM NIPLES")
        Part.objects.create(number="82-TEST", client=expected_client)

        resolved = resolve_program_client(client_reference="#N/A", part_number="82-TEST")

        self.assertEqual(resolved, expected_client)

    def test_program_table_filters_by_search_and_status(self):
        self.user.is_superuser = True
        self.user.save(update_fields=["is_superuser"])
        self.client.force_login(self.user)
        part = Part.objects.create(number="FILTER-PART")
        ProductionOrder.objects.create(
            folio="FILTER-OPEN", program="S40", part=part,
            quantity=10, remaining_quantity=10,
        )
        ProductionOrder.objects.create(
            folio="FILTER-DONE", program="S41", part=part,
            quantity=10, remaining_quantity=0, status=ProductionOrder.Status.COMPLETE,
        )

        response = self.client.get(reverse("order-list"), {"q": "FILTER", "status": "OPEN"})

        self.assertContains(response, "FILTER-OPEN")
        self.assertNotContains(response, "FILTER-DONE")

    def test_order_can_be_edited_and_recalculates_available_balance(self):
        self.user.is_superuser = True
        self.user.save(update_fields=["is_superuser"])
        self.client.force_login(self.user)
        part = Part.objects.create(number="EDIT-PART")
        order = ProductionOrder.objects.create(
            folio="EDIT-1", program="S40", part=part,
            quantity=100, remaining_quantity=60,
        )

        response = self.client.post(reverse("edit-order", args=[order.pk]), {
            "program": "S41", "part": part.pk, "quantity": 120,
            "required_date": "", "line": "L2",
        })

        self.assertRedirects(response, reverse("order-list"))
        order.refresh_from_db()
        self.assertEqual(order.program, "S41")
        self.assertEqual(order.remaining_quantity, Decimal("80"))

    def test_order_deletion_is_confirmed_and_blocked_with_production(self):
        self.user.is_superuser = True
        self.user.save(update_fields=["is_superuser"])
        self.client.force_login(self.user)
        part = Part.objects.create(number="DELETE-PART")
        removable = ProductionOrder.objects.create(
            folio="DELETE-1", program="S42", part=part,
            quantity=10, remaining_quantity=10,
        )
        response = self.client.post(reverse("delete-order", args=[removable.pk]))
        self.assertRedirects(response, reverse("order-list"))
        self.assertFalse(ProductionOrder.objects.filter(pk=removable.pk).exists())

        protected = ProductionOrder.objects.create(
            folio="DELETE-2", program="S43", part=part,
            quantity=10, remaining_quantity=0,
        )
        machine = Machine.objects.create(code="DELETE-MACHINE")
        WorkInProcess.objects.create(
            folio="DELETE-WORK", order=protected, machine=machine,
            initial_quantity=10, remaining_quantity=10, started_at=timezone.now(),
        )
        self.client.post(reverse("delete-order", args=[protected.pk]))
        self.assertTrue(ProductionOrder.objects.filter(pk=protected.pk).exists())

    def test_surplus_allocation_updates_both_balances(self):
        part = Part.objects.create(number="P-200")
        inventory = Inventory.objects.create(part=part, surplus=10, real=10)
        move_surplus(part=part, action="ALLOCATE", quantity=4, employee=self.employee,
                     program="OP-80", user=self.user)
        inventory.refresh_from_db()
        bucket = InventoryBucket.objects.get(inventory=inventory, kind="PROGRAM", name="OP-80")
        self.assertEqual(inventory.surplus, Decimal("6"))
        self.assertEqual(bucket.quantity, Decimal("4"))

    def test_surplus_cannot_become_negative(self):
        part = Part.objects.create(number="P-300")
        Inventory.objects.create(part=part, surplus=2, real=2)
        with self.assertRaises(ValidationError):
            move_surplus(part=part, action="ALLOCATE", quantity=3,
                         employee=self.employee, program="OP-90", user=self.user)
        self.assertFalse(Movement.objects.filter(part=part).exists())

    def test_material_moves_between_processes(self):
        part = Part.objects.create(number="P-400")
        inventory = Inventory.objects.create(part=part, real=12)
        cutting = Process.objects.create(code="corte", name="Corte", position=1)
        bending = Process.objects.create(code="doblez", name="Doblez", position=2)
        InventoryBucket.objects.create(inventory=inventory, kind="PROCESS", name="Corte", quantity=12)
        move_process_material(part=part, source_process=cutting, destination_process=bending,
                              program="OP-100", quantity=5, employee=self.employee, user=self.user)
        self.assertEqual(InventoryBucket.objects.get(inventory=inventory, name="Corte").quantity,
                         Decimal("7"))
        self.assertEqual(InventoryBucket.objects.get(inventory=inventory, name="Doblez").quantity,
                         Decimal("5"))
